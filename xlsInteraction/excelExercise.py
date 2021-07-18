import openpyxl

#  load workbook
inventory_file = openpyxl.load_workbook("inventory.xlsx")
print(type(inventory_file))

# load worksheet
product_list = inventory_file["prod_list"]
print(type(product_list))

"""
Task #1 - how many products we have per supplier and list the name of supplier with the respective numbers
Our output should be a dictionary
{'product1':count1,'product2':count2,'product3':count3}

Task# 2 - We need to calculate the net inventory values of all products per supplier
"""
product_per_supplier = {}
total_value_per_supplier = {}
print(type(product_per_supplier))

# calculate no of rows
print(product_list.max_row)

#  iterating over a list we need to use range() - creates a sequence of numbers
#  in range function if we don not provide any starting index it will be always start with zero
#  But zero actually gives error as there is no zero index , if we have first rows as headers,
#  then we need to pass starting index as 2
#  if 1st row contains data without header our range starts from 1 till max_row
#  Last Index is always is exclusive (not included by default) and we need to add 1 to get it fixed

for product_row in range(2, product_list.max_row + 1):
    # we need to get 4th column to get the Supplier values
    # we need to use row number and column number in cell() as it expects two params and return values according to that
    # we need to get the value of that cell other wise it will add cell object values to the list while iterating
    supplier_name = product_list.cell(product_row, 4).value
    inventory = product_list.cell(product_row, 2).value
    price = product_list.cell(product_row, 3).value

    # we are using dictionary[key] to set and dictionary.get(key) to fetch the value
    # calculation logic for counts of no of products per supplier (task #1)
    if supplier_name in product_per_supplier:
        product_per_supplier[supplier_name] = product_per_supplier.get(supplier_name) + 1
    else:
        print("adding a new supplier")
        product_per_supplier[supplier_name] = 1

    # calculation logic for inventory total value per supplier (task #2)
    if supplier_name in total_value_per_supplier:
        total_value_per_supplier[supplier_name] = total_value_per_supplier.get(supplier_name) + inventory * price
    else:
        print("adding new key for inventory values")
        total_value_per_supplier[supplier_name] = inventory * price

print(product_per_supplier)
print(total_value_per_supplier)
